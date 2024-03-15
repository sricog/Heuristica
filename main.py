from algoritmos import constructivo, grasp,grasp_ruido
from auxiliares import agregar_datos
from openpyxl import Workbook

def main():
    nsol = 1100
    k = 3
    r = 8
    wb0 = Workbook()
    wb1 = Workbook()
    wb2 = Workbook()
    for id in range(1,22):
        file_path = f'TOPinstances/TOP{id}.txt' #Si la ruta es diferente, modificarla
        rutas0,z0,t0 = constructivo(file_path)
        rutas1,z1,t1 = grasp(file_path,k,nsol)
        rutas2,z2,t2 = grasp_ruido(file_path,k,r,nsol)
        for q,wb in enumerate(list((wb0,wb1,wb2))):
            ws = wb.create_sheet(f'TOP{id}')
            if q == 0:
                agregar_datos(wb,ws,rutas0,z0,t0,f"CONSTRUCTIVO")
            elif q == 1:
                agregar_datos(wb,ws,rutas1,z1,t1,f"GRASP")
            elif q == 2:
                agregar_datos(wb,ws,rutas2,z2,t2,f"GRASP_RUIDO")

if __name__ == "__main__":
    main()
